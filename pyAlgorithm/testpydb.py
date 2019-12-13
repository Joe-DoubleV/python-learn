import pymssql
import zipfile
import openpyxl
# import connect
server = ""
user = "sa"
password = ""
database = ""
path = "C:\\Users\\0104795\\Desktop\\print\\test\\file.zip"
localpath = ".\\test\\"
srcpath = ".\\test\\fiblist.xlsx"
target = "\\pythonfile\\mocc\\pyAlgorithm\\test\\test.zip"

def getFileData(FibreNo):    
    with pymssql.connect(server,user,password,database) as conn:
        with conn.cursor(as_dict=True) as cursor:
            sql = 'select * from OTDR8KFile_H where FibreNo = \'{}\''
            cursor.execute(sql.format(FibreNo))
            print(len(cursor.fetchall()))
            for row in cursor:
                filedata = row['FileData']
                with open(localpath+FibreNo+".zip",'wb') as f:
                    f.write(filedata)

                if zipfile.is_zipfile(localpath):
                    z = zipfile.ZipFile(localpath,'r')
                    for file in z.namelist():
                        z.extract(file,'.\\test\\')
                else:
                    pass
                

def getFiberNo():
    wb = openpyxl.load_workbook(srcpath)
    ws = wb.active
    print(ws)
    # colA = ws['C']
    print(ws.max_column,ws.max_row)
    cloA = ws['A']
    for a in cloA:
        if not a.value:
            b = a
            continue

        fiberNo = a.value[0:11]
        if 11 == len(fiberNo):
            getFileData(fiberNo)
            break


        if not a.value and not b.value:
            break
        b = a
        

    
    

print(getFiberNo())
